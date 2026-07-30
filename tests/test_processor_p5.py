# tests/test_processor_p5.py
# 职责：Phase 5 端到端集成测试 — mock 模式验证双报告生成

import sys
import os
import tempfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__))))

import unittest
from src.email_connector import EmailConnector
from src.email_fetcher import EmailFetcher
from src.processor import EmailProcessor
from src.email_store import EmailStore
from tests.mock_imap import MockIMAPConnection


class TestPhase5Integration(unittest.TestCase):
    """Phase 5 集成测试：mock 端到端验证双报告生成"""

    def setUp(self):
        self.samples_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "samples"
        )
        self.config = {
            "email": {
                "server": "imap.qiye.aliyun.com",
                "port": 993,
                "account": "test@mock.com",
                "password": "testpass",
                "auth_type": "password",
            },
            "keywords": {
                "invoice_body": ["发票", "开票"],
                "invoice_table": "开票申请",
                "urgent": ["加急"],
                "table_parser": {
                    "mode_priority": ["key_value", "column_index"],
                    "field_mapping": {
                        "amount": ["开票金额", "金额", "实付金额"],
                        "order_id": ["订单号", "订单编号", "主订单ID"],
                        "note": ["备注"],
                    },
                    "column_indices": {
                        "amount": 8,
                        "order_id": 10,
                        "note": 14,
                    },
                },
            },
            "schedule": {"interval": "1h"},
            "output": {"dir": "", "filename_pattern": ""},
            "order_no": {"valid_lengths": [12, 16]},
            "dedup": {"check_history": True},
        }
        self.tmpdir = tempfile.mkdtemp()
        self.store_path = os.path.join(self.tmpdir, "test_processed.json")
        self.store = EmailStore(self.store_path)

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_processor(self):
        """创建带 mock 的完整处理链"""
        # 如果未设置 output.dir，使用 tmpdir
        if not self.config["output"]["dir"]:
            self.config["output"]["dir"] = self.tmpdir
        mock_imap = MockIMAPConnection(
            self.samples_dir, "test@mock.com", "testpass"
        )
        connector = EmailConnector(self.config, mock_imap=mock_imap)
        connector.connect()
        fetcher = EmailFetcher(connector)
        processor = EmailProcessor(self.config, connector, fetcher, self.store)
        return processor

    def test_reports_generated_after_run(self):
        """运行后报告文件应生成到输出目录"""
        processor = self._make_processor()
        result = processor.run()
        self.assertTrue(os.path.exists(result.report_md),
                        f".md 报告不存在: {result.report_md}")
        self.assertTrue(os.path.exists(result.report_xlsx),
                        f".xlsx 报告不存在: {result.report_xlsx}")

    def test_md_report_is_valid_markdown(self):
        """.md 报告应为有效 UTF-8 文本，含 Markdown 表格语法"""
        processor = self._make_processor()
        result = processor.run()
        with open(result.report_md, "r", encoding="utf-8") as f:
            content = f.read()
        self.assertIn("# 发票邮件处理报告", content)
        self.assertIn("|---", content)  # Markdown 表格分隔符
        self.assertIn("| 订单号 |", content)  # 订单表头

    def test_xlsx_has_5_sheets(self):
        """.xlsx 报告应有 5 个 Sheet"""
        from openpyxl import load_workbook
        processor = self._make_processor()
        result = processor.run()
        wb = load_workbook(result.report_xlsx)
        self.assertEqual(len(wb.sheetnames), 5)
        self.assertIn("汇总", wb.sheetnames)
        self.assertIn("加急订单", wb.sheetnames)
        self.assertIn("正常订单", wb.sheetnames)
        self.assertIn("订单号异常", wb.sheetnames)
        self.assertIn("疑似不确定邮件", wb.sheetnames)

    def test_urgent_sheet_has_yellow_fill(self):
        """加急 Sheet 应有黄色背景"""
        from openpyxl import load_workbook
        processor = self._make_processor()
        result = processor.run()
        wb = load_workbook(result.report_xlsx)
        ws = wb["加急订单"]
        # 数据行（行索引 2+）应有黄色填充
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value:
                    fill = cell.fill
                    self.assertEqual(
                        fill.start_color.rgb if fill.start_color else None,
                        "00FFFF00",
                        f"单元格 {cell.coordinate} 应为黄色背景"
                    )

    def test_output_dir_auto_created(self):
        """输出目录不存在时应自动创建"""
        new_dir = os.path.join(tempfile.gettempdir(), "_test_invoice_report_phase5")
        if os.path.exists(new_dir):
            import shutil
            shutil.rmtree(new_dir)
        self.config["output"]["dir"] = new_dir
        processor = self._make_processor()
        result = processor.run()
        self.assertTrue(os.path.exists(new_dir))
        self.assertTrue(os.path.exists(result.report_md))

    def test_result_has_report_paths(self):
        """ProcessingResult 应包含报告路径"""
        processor = self._make_processor()
        result = processor.run()
        self.assertNotEqual(result.report_md, "")
        self.assertNotEqual(result.report_xlsx, "")
        self.assertIn("发票邮件报告", result.report_md)
        self.assertIn("发票邮件报告", result.report_xlsx)

    def test_md_contains_required_sections(self):
        """.md 应包含头部统计、加急订单和异常区"""
        processor = self._make_processor()
        result = processor.run()
        with open(result.report_md, "r", encoding="utf-8") as f:
            content = f.read()
        self.assertIn("## 加急订单", content)
        self.assertIn("## 处理统计", content)
        self.assertIn("**运行时间**", content)

    def test_report_filename_matches_spec(self):
        """报告文件名格式匹配 {YYYY.M.D}_{HH-mm}_发票邮件报告.{md|xlsx}"""
        import re
        processor = self._make_processor()
        result = processor.run()
        md_basename = os.path.basename(result.report_md)
        xlsx_basename = os.path.basename(result.report_xlsx)
        pattern = r"^\d{4}\.\d{1,2}\.\d{1,2}_\d{2}-\d{2}_发票邮件报告\.(md|xlsx)$"
        self.assertRegex(md_basename, pattern)
        self.assertRegex(xlsx_basename, pattern)


if __name__ == "__main__":
    unittest.main()
