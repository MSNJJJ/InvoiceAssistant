# tests/test_report_generator.py
# 职责：Phase 5 报告生成器单元测试 — MD + XLSX

import sys
import os
import tempfile
import re

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(__file__))))

import unittest
from datetime import datetime
from src.report_generator import (
    ReportData, _build_report_data, _generate_md, _generate_xlsx,
    generate_report,
)
from src.order_validator import ValidatedOrder


class TestReportDataBuilding(unittest.TestCase):
    """测试 ReportData 聚合构建"""

    def setUp(self):
        # 最小的 mock processor
        class MockProcessor:
            def get_validated_orders(self):
                return self._orders
            def get_collected_uncertain(self):
                return self._uncertain

        # 最小的 mock result
        class MockResult:
            total_unread = 12
            invoice_count = 5
            urgent_count = 2
            uncertain_count = 1

        self.MockProcessor = MockProcessor
        self.MockResult = MockResult

    def _make_order(self, quadrant, is_valid=True, urgent=False):
        return ValidatedOrder(
            order_id_original="ORD-9000000784169034",
            amount_raw="3904元",
            note="测试备注",
            order_id_cleaned="9000000784169034",
            is_valid=is_valid,
            validation_reason="valid" if is_valid else "too_short(5)",
            is_urgent=urgent,
            quadrant=quadrant,
            message_subject="开票申请",
            message_sender="test@xx.com",
            message_date="2026-07-29 15:22",
            message_id="msg001",
        )

    def test_urgent_valid_goes_to_urgent(self):
        """urgent_valid 应进入加急表"""
        proc = self.MockProcessor()
        proc._orders = [self._make_order("urgent_valid")]
        proc._uncertain = []
        data = _build_report_data(proc, self.MockResult())
        self.assertEqual(len(data.urgent_orders), 1)
        self.assertEqual(len(data.normal_orders), 0)

    def test_normal_valid_goes_to_normal(self):
        """normal_valid 应进入正常表"""
        proc = self.MockProcessor()
        proc._orders = [self._make_order("normal_valid")]
        proc._uncertain = []
        data = _build_report_data(proc, self.MockResult())
        self.assertEqual(len(data.normal_orders), 1)
        self.assertEqual(len(data.urgent_orders), 0)

    def test_urgent_invalid_goes_to_both(self):
        """urgent_invalid 应进入加急表 + 异常表"""
        proc = self.MockProcessor()
        proc._orders = [self._make_order("urgent_invalid", is_valid=False, urgent=True)]
        proc._uncertain = []
        data = _build_report_data(proc, self.MockResult())
        self.assertEqual(len(data.urgent_orders), 1)
        self.assertEqual(len(data.invalid_orders), 1)

    def test_normal_invalid_goes_to_invalid(self):
        """normal_invalid 应进入异常表"""
        proc = self.MockProcessor()
        proc._orders = [self._make_order("normal_invalid", is_valid=False)]
        proc._uncertain = []
        data = _build_report_data(proc, self.MockResult())
        self.assertEqual(len(data.invalid_orders), 1)
        self.assertEqual(len(data.urgent_orders), 0)

    def test_statistics_carried_over(self):
        """统计数据应从 result 正确传递"""
        proc = self.MockProcessor()
        proc._orders = []
        proc._uncertain = []
        result = self.MockResult()
        data = _build_report_data(proc, result)
        self.assertEqual(data.total_unread, 12)
        self.assertEqual(data.invoice_count, 5)
        self.assertEqual(data.urgent_count, 2)
        self.assertEqual(data.uncertain_count, 1)


class TestMDFormatting(unittest.TestCase):
    """测试 MD 报告格式"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_empty_data(self):
        return ReportData(
            run_time="2026-07-29 17:39:53",
            total_unread=0, invoice_count=0,
            urgent_count=0, uncertain_count=0,
            urgent_orders=[], normal_orders=[], invalid_orders=[],
            uncertain_entries=[],
        )

    def _make_sample_data(self):
        order_urgent = ValidatedOrder(
            order_id_original="ORD-9000000784169034",
            amount_raw="3904元", note="加急处理",
            order_id_cleaned="9000000784169034",
            is_valid=True, validation_reason="valid",
            is_urgent=True, quadrant="urgent_valid",
            message_subject="开票申请", message_sender="zzf@xx.com",
            message_date="2026-07-29 15:22", message_id="m1",
        )
        order_normal = ValidatedOrder(
            order_id_original="9000000782190489",
            amount_raw="1880元", note="正常订单",
            order_id_cleaned="9000000782190489",
            is_valid=True, validation_reason="valid",
            is_urgent=False, quadrant="normal_valid",
            message_subject="开票申请", message_sender="edna@xx.com",
            message_date="2026-07-29 14:10", message_id="m2",
        )
        order_invalid = ValidatedOrder(
            order_id_original="12345",
            amount_raw="100元", note="测试",
            order_id_cleaned="12345",
            is_valid=False, validation_reason="too_short(5)",
            is_urgent=False, quadrant="normal_invalid",
            message_subject="开票申请", message_sender="test@xx.com",
            message_date="2026-07-29 13:00", message_id="m3",
        )
        return ReportData(
            run_time="2026-07-29 17:39:53",
            total_unread=12, invoice_count=5,
            urgent_count=1, uncertain_count=1,
            urgent_orders=[order_urgent],
            normal_orders=[order_normal],
            invalid_orders=[order_invalid],
            uncertain_entries=[{
                "message": type("Msg", (), {"subject": "不确定邮件", "sender": "unk@xx.com", "date": "2026-07-29 12:00"})(),
                "classification": type("Cls", (), {"reasons": ["正文过短", "无表格"]})(),
            }],
        )

    def test_empty_data_generates_valid_md(self):
        """空数据应生成有效 .md 文件"""
        path = os.path.join(self.tmpdir, "empty.md")
        _generate_md(self._make_empty_data(), path)
        self.assertTrue(os.path.exists(path))
        content = open(path, encoding="utf-8").read()
        self.assertIn("发票邮件处理报告", content)
        self.assertIn("0", content)

    def test_urgent_section_first_with_bold(self):
        """加急区应在最前，订单号被 ** 包裹"""
        path = os.path.join(self.tmpdir, "urgent.md")
        _generate_md(self._make_sample_data(), path)
        content = open(path, encoding="utf-8").read()
        # 加急区在正常区之前
        urgent_pos = content.index("加急订单")
        normal_pos = content.index("正常订单")
        self.assertLess(urgent_pos, normal_pos)
        # 订单号加粗
        self.assertIn("**9000000784169034**", content)

    def test_normal_no_bold(self):
        """正常订单号无加粗"""
        path = os.path.join(self.tmpdir, "normal.md")
        _generate_md(self._make_sample_data(), path)
        content = open(path, encoding="utf-8").read()
        # 正常订单号不加粗
        normal_section = content.split("## 正常订单")[1].split("## ")[0]
        self.assertIn("9000000782190489", normal_section)
        self.assertNotIn("**9000000782190489**", normal_section)

    def test_invalid_has_8_columns(self):
        """异常区应有 8 列：订单号原文|清洗后数字|异常原因|开票金额|备注|来源邮件主题|发件人|邮件时间"""
        path = os.path.join(self.tmpdir, "invalid.md")
        _generate_md(self._make_sample_data(), path)
        content = open(path, encoding="utf-8").read()
        # 检查异常表头有 8 列
        header_line = [l for l in content.split("\n") if "订单号原文" in l][0]
        cols = header_line.split("|")
        # 去掉首尾空，应该有 10 个管道符分隔 = 9 段，但首尾 | 是空 → 8 列
        self.assertEqual(len([c for c in cols if c.strip()]), 8, f"异常区应有 8 列, 实际列: {cols}")

    def test_filename_format(self):
        """文件名匹配 {YYYY.M.D}_{HH-mm}_发票邮件报告.md"""
        from src.report_generator import generate_report

        # Mock processor/result
        class MockP:
            def get_validated_orders(self): return []
            def get_collected_uncertain(self): return []

        class MockR:
            total_unread = 0
            invoice_count = 0
            urgent_count = 0
            uncertain_count = 0

        md_path, xlsx_path = generate_report(MockP(), MockR(), self.tmpdir)
        md_basename = os.path.basename(md_path)
        xlsx_basename = os.path.basename(xlsx_path)
        # 命名格式：日期_时间_发票邮件报告.{md|xlsx}
        self.assertTrue(md_basename.endswith("_发票邮件报告.md"), md_basename)
        self.assertTrue(xlsx_basename.endswith("_发票邮件报告.xlsx"), xlsx_basename)
        # 日期部分应匹配 YYYY.M.D 模式
        date_part = md_basename.split("_")[0]
        self.assertRegex(date_part, r"^\d{4}\.\d{1,2}\.\d{1,2}$")

    def test_uncertain_section_columns(self):
        """疑似不确定区应有 4 列"""
        path = os.path.join(self.tmpdir, "uncertain.md")
        _generate_md(self._make_sample_data(), path)
        content = open(path, encoding="utf-8").read()
        self.assertIn("疑似不确定邮件", content)
        # 找到分隔行（|---|...），统计列数
        uncertain_section = content.split("## 疑似不确定邮件")[1].split("##")[0]
        separator_line = [l for l in uncertain_section.split("\n") if "---" in l][0]
        cols = [c for c in separator_line.split("|") if c.strip()]
        self.assertEqual(len(cols), 4)

    def test_md_table_separator(self):
        """MD 文件应含 Markdown 表格分隔符"""
        path = os.path.join(self.tmpdir, "tables.md")
        _generate_md(self._make_sample_data(), path)
        content = open(path, encoding="utf-8").read()
        self.assertIn("|---", content)

    def test_column_names_correct(self):
        """各表列名应与 REQUIREMENTS.md 一致"""
        path = os.path.join(self.tmpdir, "columns.md")
        _generate_md(self._make_sample_data(), path)
        content = open(path, encoding="utf-8").read()
        # 加急/正常订单表头
        self.assertIn("| 订单号 | 开票金额 | 备注 | 来源邮件主题 | 发件人 | 邮件时间 |", content)
        # 异常表头
        self.assertIn("| 订单号原文 | 清洗后数字 | 异常原因 | 开票金额 | 备注 | 来源邮件主题 | 发件人 | 邮件时间 |", content)
        # 疑似表头
        self.assertIn("| 邮件主题 | 发件人 | 邮件时间 | 不确定原因 |", content)

    def test_run_time_in_header(self):
        """报告头部应含运行时间"""
        path = os.path.join(self.tmpdir, "time.md")
        _generate_md(self._make_sample_data(), path)
        content = open(path, encoding="utf-8").read()
        self.assertIn("**运行时间**：2026-07-29 17:39:53", content)


class TestXLSXFormatting(unittest.TestCase):
    """测试 XLSX 报告格式"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def tearDown(self):
        import shutil
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _make_empty_data(self):
        return ReportData(
            run_time="2026-07-29 17:39:53",
            total_unread=0, invoice_count=0,
            urgent_count=0, uncertain_count=0,
            urgent_orders=[], normal_orders=[], invalid_orders=[],
            uncertain_entries=[],
        )

    def _make_sample_data(self):
        order_urgent = ValidatedOrder(
            order_id_original="ORD-9000000784169034",
            amount_raw="3904元", note="加急处理",
            order_id_cleaned="9000000784169034",
            is_valid=True, validation_reason="valid",
            is_urgent=True, quadrant="urgent_valid",
            message_subject="开票申请", message_sender="zzf@xx.com",
            message_date="2026-07-29 15:22", message_id="m1",
        )
        order_normal = ValidatedOrder(
            order_id_original="9000000782190489",
            amount_raw="1880元", note="正常订单",
            order_id_cleaned="9000000782190489",
            is_valid=True, validation_reason="valid",
            is_urgent=False, quadrant="normal_valid",
            message_subject="开票申请", message_sender="edna@xx.com",
            message_date="2026-07-29 14:10", message_id="m2",
        )
        order_invalid = ValidatedOrder(
            order_id_original="12345",
            amount_raw="100元", note="测试",
            order_id_cleaned="12345",
            is_valid=False, validation_reason="too_short(5)",
            is_urgent=False, quadrant="normal_invalid",
            message_subject="开票申请", message_sender="test@xx.com",
            message_date="2026-07-29 13:00", message_id="m3",
        )
        return ReportData(
            run_time="2026-07-29 17:39:53",
            total_unread=12, invoice_count=5,
            urgent_count=1, uncertain_count=1,
            urgent_orders=[order_urgent],
            normal_orders=[order_normal],
            invalid_orders=[order_invalid],
            uncertain_entries=[{
                "message": type("Msg", (), {"subject": "不确定邮件", "sender": "unk@xx.com", "date": "2026-07-29 12:00"})(),
                "classification": type("Cls", (), {"reasons": ["正文过短", "无表格"]})(),
            }],
        )

    def test_empty_data_generates_valid_xlsx(self):
        """空数据应生成有效 .xlsx 文件，含 5 个 Sheet"""
        from openpyxl import load_workbook
        path = os.path.join(self.tmpdir, "empty.xlsx")
        _generate_xlsx(self._make_empty_data(), path)
        self.assertTrue(os.path.exists(path))
        wb = load_workbook(path)
        self.assertEqual(len(wb.sheetnames), 5)

    def test_sheet_order(self):
        """Sheet 顺序应为：汇总/加急/正常/异常/疑似"""
        from openpyxl import load_workbook
        path = os.path.join(self.tmpdir, "order.xlsx")
        _generate_xlsx(self._make_sample_data(), path)
        wb = load_workbook(path)
        self.assertEqual(wb.sheetnames, ["汇总", "加急订单", "正常订单", "订单号异常", "疑似不确定邮件"])

    def test_urgent_sheet_has_yellow_fill(self):
        """加急 Sheet 应有黄色背景"""
        from openpyxl import load_workbook
        path = os.path.join(self.tmpdir, "urgent.xlsx")
        _generate_xlsx(self._make_sample_data(), path)
        wb = load_workbook(path)
        ws = wb["加急订单"]
        # 数据行（行索引 2+）应有黄色填充
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            has_data = any(cell.value for cell in row)
            if has_data:
                for cell in row:
                    fill = cell.fill
                    if fill and fill.start_color:
                        self.assertEqual(
                            fill.start_color.rgb if fill.start_color else None,
                            "00FFFF00",
                            f"单元格 {cell.coordinate} 应为黄色背景"
                        )

    def test_summary_sheet_statistics(self):
        """汇总 Sheet 统计应正确"""
        from openpyxl import load_workbook
        path = os.path.join(self.tmpdir, "summary.xlsx")
        _generate_xlsx(self._make_sample_data(), path)
        wb = load_workbook(path)
        ws = wb["汇总"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        data = {row[0]: row[1] for row in rows if row[0]}
        self.assertEqual(data["处理未读邮件数"], 12)
        self.assertEqual(data["开票邮件数"], 5)
        self.assertEqual(data["加急订单数"], 1)

    def test_sheet_columns_correct(self):
        """各 Sheet 列名应正确"""
        from openpyxl import load_workbook
        path = os.path.join(self.tmpdir, "columns.xlsx")
        _generate_xlsx(self._make_sample_data(), path)
        wb = load_workbook(path)
        # 加急 Sheet
        ws2 = wb["加急订单"]
        headers2 = [cell.value for cell in ws2[1]]
        self.assertEqual(headers2, ["订单号", "开票金额", "备注", "来源邮件主题", "发件人", "邮件时间"])
        # 异常 Sheet
        ws4 = wb["订单号异常"]
        headers4 = [cell.value for cell in ws4[1]]
        self.assertEqual(headers4, ["订单号原文", "清洗后数字", "异常原因", "开票金额", "备注", "来源邮件主题", "发件人", "邮件时间"])
        # 疑似 Sheet
        ws5 = wb["疑似不确定邮件"]
        headers5 = [cell.value for cell in ws5[1]]
        self.assertEqual(headers5, ["邮件主题", "发件人", "邮件时间", "不确定原因"])


if __name__ == "__main__":
    unittest.main()
